"""
Fetches the top 5 most recent "scam news Malaysia" articles via Google News RSS,
extracts full article content using trafilatura (with readability as fallback),
and saves to Excel.

Requirements:
    pip install feedparser trafilatura readability-lxml requests openpyxl
"""

import feedparser
import requests
import trafilatura
from readability import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time

# ── Config ────────────────────────────────────────────────────────────────────
QUERY        = "scam news Malaysia"
TOP_N        = 5
OUTPUT_FILE  = "scam_news_malaysia.xlsx"
HEADERS      = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

# ── Step 1: Fetch top-N articles from Google News RSS ─────────────────────────
def fetch_news_rss(query: str, n: int) -> list[dict]:
    rss_url = (
        f"https://news.google.com/rss/search"
        f"?q={requests.utils.quote(query)}&hl=en-MY&gl=MY&ceid=MY:en"
    )
    feed = feedparser.parse(rss_url)
    entries = feed.entries[:n]

    results = []
    for entry in entries:
        published = entry.get("published", "")
        try:
            pub_dt = datetime(*entry.published_parsed[:6])
        except Exception:
            pub_dt = None

        results.append({
            "title":     entry.get("title", "N/A"),
            "url":       entry.get("link", ""),
            "source":    entry.get("source", {}).get("title", "N/A"),
            "published": pub_dt,
        })
    return results

# ── Step 2: Extract full article content ──────────────────────────────────────
def extract_content(url: str) -> str:
    try:
        html = requests.get(url, headers=HEADERS, timeout=15).text

        # Primary: trafilatura (best-in-class news extractor)
        text = trafilatura.extract(
            html,
            include_comments=False,
            include_tables=False,
            no_fallback=False,
            favor_recall=True,  # prioritise getting more content
        )
        if text and len(text.strip()) > 100:
            return text.strip()

        # Fallback: readability-lxml (strips boilerplate, keeps main body)
        doc  = Document(html)
        body = doc.summary(html_partial=True)
        # Strip HTML tags from readability output
        from html.parser import HTMLParser
        class _Strip(HTMLParser):
            def __init__(self):
                super().__init__()
                self.chunks = []
            def handle_data(self, data):
                self.chunks.append(data)
        parser = _Strip()
        parser.feed(body)
        plain = " ".join(parser.chunks).strip()
        if plain and len(plain) > 100:
            return plain

        return "[Content could not be extracted — site may require JS or login]"

    except Exception as e:
        return f"[Error: {e}]"

# ── Step 3: Write to Excel ─────────────────────────────────────────────────────
def write_excel(articles: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scam News Malaysia"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color="1F4E79")
    cell_font      = Font(name="Arial", size=10)
    wrap_align     = Alignment(wrap_text=True, vertical="top")
    center_align   = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    alt_fill       = PatternFill("solid", start_color="D6E4F0")

    # ── Column headers ────────────────────────────────────────────────────
    columns = ["#", "Title", "Source", "Published Date", "URL", "Article Content"]
    col_widths = [4, 40, 20, 18, 35, 80]

    for col_idx, (header, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, art in enumerate(articles, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        values = [
            row_idx - 1,
            art["title"],
            art["source"],
            art["published"].strftime("%d %b %Y %H:%M") if art["published"] else "N/A",
            art["url"],
            art["content"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = cell_font
            cell.border = thin_border
            cell.fill   = fill
            cell.alignment = center_align if col_idx == 1 else wrap_align

        # Taller row for content
        ws.row_dimensions[row_idx].height = 120

    # ── Freeze header row ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Metadata sheet ────────────────────────────────────────────────────
    meta = wb.create_sheet("Metadata")
    meta_font  = Font(name="Arial", size=10)
    meta_bold  = Font(name="Arial", size=10, bold=True)
    meta_data = [
        ("Search Query",   QUERY),
        ("Results Fetched", len(articles)),
        ("Source",         "Google News RSS"),
        ("Generated On",   datetime.now().strftime("%d %b %Y %H:%M:%S")),
    ]
    for r, (key, val) in enumerate(meta_data, start=1):
        meta.cell(row=r, column=1, value=key).font  = meta_bold
        meta.cell(row=r, column=2, value=val).font  = meta_font
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 45

    wb.save(output_path)
    print(f"\n✅ Saved → {output_path}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f'🔍 Searching Google News RSS for: "{QUERY}"')
    articles = fetch_news_rss(QUERY, TOP_N)

    if not articles:
        print("❌ No results found. Check your internet connection.")
        return

    print(f"📰 Found {len(articles)} articles. Extracting content...\n")
    for i, art in enumerate(articles, start=1):
        print(f"  [{i}/{len(articles)}] {art['title'][:70]}...")
        art["content"] = extract_content(art["url"])
        time.sleep(1)  # polite crawl delay

    write_excel(articles, OUTPUT_FILE)

if __name__ == "__main__":
    main()